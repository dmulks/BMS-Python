"""
Excel Import/Export Service using pandas and openpyxl.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from datetime import date, datetime
from decimal import Decimal
from typing import List, Dict
import os

from app.models.bond import BondPurchase, BondType
from app.models.payment import CouponPayment
from app.models.user import User
from app.models.balance import MemberBalance, MonthlySummary


class ExcelService:
    """Service for Excel import and export operations."""

    @staticmethod
    def export_monthly_summary(db: Session, month: date) -> str:
        """
        Export monthly summary to Excel.

        Args:
            db: Database session
            month: Month to export

        Returns:
            Path to generated Excel file
        """
        # Get summary data
        summary = db.query(MonthlySummary).filter(
            MonthlySummary.summary_month == month
        ).first()

        if not summary:
            raise ValueError(f"No summary found for {month}")

        # Get member balances
        balances = db.query(MemberBalance).filter(
            MemberBalance.balance_date == month
        ).all()

        # Create workbook
        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Monthly Summary"

        # Header styling
        header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws_summary['A1'] = "BOND COOPERATIVE SOCIETY"
        ws_summary['A1'].font = Font(bold=True, size=16, color="1a237e")
        ws_summary['A2'] = f"Monthly Summary - {month.strftime('%B %Y')}"
        ws_summary['A2'].font = Font(bold=True, size=14)

        # Summary data
        row = 4
        summary_data = [
            ("Total Bond Shares", float(summary.total_bond_shares)),
            ("Total Face Value", float(summary.total_face_value)),
            ("Total Purchases", float(summary.total_purchases)),
            ("Total Gross Coupons", float(summary.total_gross_coupons)),
            ("Total Withholding Tax", float(summary.total_withholding_tax)),
            ("Total BOZ Fees", float(summary.total_boz_fees)),
            ("Total Co-op Fees", float(summary.total_coop_fees)),
            ("Total Net Payments", float(summary.total_net_payments)),
            ("Net Cooperative Income", float(summary.net_cooperative_income)),
            ("Active Members Count", summary.active_members_count),
            ("New Purchases Count", summary.new_purchases_count),
            ("Matured Bonds Count", summary.matured_bonds_count),
        ]

        for label, value in summary_data:
            ws_summary[f'A{row}'] = label
            ws_summary[f'A{row}'].font = Font(bold=True)
            ws_summary[f'B{row}'] = value
            if isinstance(value, float):
                ws_summary[f'B{row}'].number_format = '#,##0.00'
            row += 1

        # Member Balances sheet
        ws_balances = wb.create_sheet("Member Balances")

        # Headers
        headers = [
            'Member ID', 'Member Name', 'Bond Type', 'Opening Balance',
            'Purchases', 'Payments Received', 'Closing Balance',
            'Total Shares', 'Total Face Value', '% Share'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws_balances.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Data
        row = 2
        for balance in balances:
            user = db.query(User).filter(User.user_id == balance.user_id).first()
            bond_type = db.query(BondType).filter(BondType.bond_type_id == balance.bond_type_id).first()

            ws_balances.cell(row=row, column=1, value=f"M{user.user_id:04d}")
            ws_balances.cell(row=row, column=2, value=f"{user.first_name} {user.last_name}")
            ws_balances.cell(row=row, column=3, value=bond_type.bond_name)
            ws_balances.cell(row=row, column=4, value=float(balance.opening_balance))
            ws_balances.cell(row=row, column=5, value=float(balance.purchases_month))
            ws_balances.cell(row=row, column=6, value=float(balance.payments_received))
            ws_balances.cell(row=row, column=7, value=float(balance.closing_balance))
            ws_balances.cell(row=row, column=8, value=float(balance.total_bond_shares))
            ws_balances.cell(row=row, column=9, value=float(balance.total_face_value))
            ws_balances.cell(row=row, column=10, value=float(balance.percentage_share))

            # Number formatting
            for col in range(4, 10):
                ws_balances.cell(row=row, column=col).number_format = '#,##0.00'
            ws_balances.cell(row=row, column=10).number_format = '0.00000%'

            row += 1

        # Auto-adjust column widths
        for ws in [ws_summary, ws_balances]:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

        # Save file
        temp_dir = os.path.join(os.path.dirname(__file__), '../../temp')
        os.makedirs(temp_dir, exist_ok=True)

        filename = f"monthly_summary_{month.strftime('%Y_%m')}.xlsx"
        filepath = os.path.join(temp_dir, filename)

        wb.save(filepath)
        return filepath

    @staticmethod
    def export_payment_register(db: Session, start_date: date, end_date: date) -> str:
        """
        Export payment register to Excel.

        Args:
            db: Database session
            start_date: Start date
            end_date: End date

        Returns:
            Path to generated Excel file
        """
        # Get payments
        payments = db.query(CouponPayment).filter(
            CouponPayment.payment_date >= start_date,
            CouponPayment.payment_date <= end_date
        ).order_by(CouponPayment.payment_date).all()

        # Create DataFrame
        data = []
        for payment in payments:
            user = db.query(User).filter(User.user_id == payment.user_id).first()
            bond_purchase = db.query(BondPurchase).filter(
                BondPurchase.purchase_id == payment.purchase_id
            ).first()

            data.append({
                'Payment Date': payment.payment_date,
                'Payment Reference': payment.payment_reference,
                'Member ID': f"M{user.user_id:04d}",
                'Member Name': f"{user.first_name} {user.last_name}",
                'Payment Type': payment.payment_type.value,
                'Period Start': payment.payment_period_start,
                'Period End': payment.payment_period_end,
                'Calendar Days': payment.calendar_days,
                'Gross Coupon': float(payment.gross_coupon_amount),
                'Withholding Tax': float(payment.withholding_tax),
                'BOZ Fees': float(payment.boz_fees),
                'Co-op Fees': float(payment.coop_fees),
                'Net Payment': float(payment.net_payment_amount),
                'Status': payment.payment_status.value,
            })

        df = pd.DataFrame(data)

        # Create workbook
        temp_dir = os.path.join(os.path.dirname(__file__), '../../temp')
        os.makedirs(temp_dir, exist_ok=True)

        filename = f"payment_register_{start_date}_{end_date}.xlsx"
        filepath = os.path.join(temp_dir, filename)

        # Write to Excel with formatting
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Payment Register', index=False)

            workbook = writer.book
            worksheet = writer.sheets['Payment Register']

            # Header styling
            header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for col_num, value in enumerate(df.columns.values, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Number formatting
            for row in range(2, len(df) + 2):
                for col in [9, 10, 11, 12, 13]:  # Amount columns
                    worksheet.cell(row=row, column=col).number_format = '#,##0.00'

        return filepath

    @staticmethod
    def import_bond_purchases(db: Session, file_path: str, imported_by: int) -> Dict:
        """
        Import bond purchases from Excel file.

        Args:
            db: Database session
            file_path: Path to Excel file
            imported_by: User ID performing the import

        Returns:
            Dictionary with import results
        """
        try:
            # Read Excel file
            df = pd.read_excel(file_path)

            results = {
                'success': [],
                'errors': [],
                'total': len(df)
            }

            for idx, row in df.iterrows():
                try:
                    # Validate required fields
                    required_fields = ['email', 'bond_shares', 'purchase_date', 'bond_type']
                    for field in required_fields:
                        if field not in row or pd.isna(row[field]):
                            raise ValueError(f"Missing required field: {field}")

                    # Find or create user
                    user = db.query(User).filter(User.email == row['email']).first()
                    if not user:
                        raise ValueError(f"User not found with email: {row['email']}")

                    # Find bond type
                    bond_type = db.query(BondType).filter(
                        BondType.bond_name == row['bond_type']
                    ).first()
                    if not bond_type:
                        raise ValueError(f"Bond type not found: {row['bond_type']}")

                    # Calculate values using Bond Calculator
                    from app.services.bond_calculator import BondCalculator

                    calc_results = BondCalculator.calculate_purchase_breakdown(
                        bond_shares=Decimal(str(row['bond_shares'])),
                        purchase_date=pd.to_datetime(row['purchase_date']).date(),
                        maturity_years=bond_type.maturity_period_years,
                        discount_rate=Decimal(str(row.get('discount_rate', 0.10)))
                    )

                    # Create bond purchase
                    purchase = BondPurchase(
                        user_id=user.user_id,
                        bond_type_id=bond_type.bond_type_id,
                        purchase_date=pd.to_datetime(row['purchase_date']).date(),
                        purchase_month=pd.to_datetime(row['purchase_date']).date().replace(day=1),
                        bond_shares=Decimal(str(row['bond_shares'])),
                        face_value=calc_results['face_value'],
                        discount_value=calc_results['discount_value'],
                        coop_discount_fee=calc_results['coop_discount_fee'],
                        net_discount_value=calc_results['net_discount_value'],
                        purchase_price=calc_results['purchase_price'],
                        maturity_date=calc_results['maturity_date'],
                        transaction_reference=f"IMP{datetime.now().strftime('%Y%m%d')}{idx:04d}",
                        notes=row.get('notes', None)
                    )

                    db.add(purchase)
                    db.commit()

                    results['success'].append({
                        'row': idx + 2,  # Excel row number
                        'email': row['email'],
                        'bond_shares': row['bond_shares']
                    })

                except Exception as e:
                    db.rollback()
                    results['errors'].append({
                        'row': idx + 2,
                        'error': str(e),
                        'data': row.to_dict()
                    })

            return results

        except Exception as e:
            raise ValueError(f"Error reading Excel file: {str(e)}")
